#!/usr/bin/env python3
"""
Erstellt FFH-LB12_Inventar_DB_v9.xlsx aus LOKAL_Geräteliste LB 12.xlsx.

Mapping der Quellspalten auf die Datenbank-Felder (articles-Tabelle):
  ID                          -> inventoryNumber
  ID MP FEUER                 -> mpFeuerInventoryNumber
  Bezeichnung LB 12           -> name
  Standort                    -> warehouse
  Geräteart                   -> deviceClass
  Seriennummer                -> serialNumber
  Inventarnummer Gemeinde     -> communityInventoryNumber
  Hersteller-Bezeichnung      -> specification
  Hersteller                  -> manufacturer
  Herstellungsdatum           -> manufacturingDate
  Datum Indienststellung       -> commissionedDate
  Prüfintervall Sicht+Funktion -> inspectionInterval
  Aussonderungsfrist           -> retirementPeriodMonths
  Datum Außerdienststellung    -> decommissionedDate
  Rechtsgrundlagen             -> pruefgrundsaetze (extra Spalte)
  Bemerkungen                  -> description

Zusätzlich werden erzeugt:
  designationLB               -> immer "LB12"
  deviceSubclass              -> abgeleitet aus deviceClass + Artikelname
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import re

# Standort -> Warehouse Name Mapping
STANDORT_MAP = {
    "12-45": "HLF 12/45",
    "12-60": "RW 12/60",
    "gh": "Gerätehalle",
    "werkstatt": "Werkstatt",
}

# Geräteart (Quelle) -> DeviceClass Name (DB) Normalisierung
DEVICE_CLASS_NORMALIZE = {
    "schutzkleidung- und schutzgerät": "Schutzkleidung und Schutzgerät",
    "schutzkleidung und schutzgerät": "Schutzkleidung und Schutzgerät",
    "löschgerät": "Löschgerät",
    "schläuche, armaturen, zubehör": "Schläuche, Armaturen und Zubehör",
    "schläuche, armaturen und zubehör": "Schläuche, Armaturen und Zubehör",
    "rettungsgerät": "Rettungsgerät",
    "sanitäts- und wiederlebungsgerät": "Sanitäts- und Wiederbelebungsgerät",
    "sanitäts- und wiederbelebungsgerät": "Sanitäts- und Wiederbelebungsgerät",
    "beleuchtungs-, signal- und fernmeldegerät": "Beleuchtungs-, Signal- und Fernmeldegerät",
    "arbeitsgerät": "Arbeitsgerät",
    "handwerkzeug und messgerät": "Handwerkzeug und Messgerät",
    "sondergerät": "Sondergerät",
    "pumpen": "Pumpen",
    "atemschutz": "Atemschutz",
}

# Prüfintervall Text -> Monate
INTERVAL_MAP = {
    "1 monat": 1,
    "4 monate": 4,
    "6 monate": 6,
    "12 monate": 12,
    "24 monate": 24,
    "36 monate": 36,
    "entfällt": None,
}

# Aussonderungsfrist Text -> Monate
RETIREMENT_MAP = {
    "1 jahr": 12,
    "5 jahre": 60,
    "8 jahre": 96,
    "10 jahre": 120,
    "unbegrenzt": None,
}

# Unterklassen-Zuordnung basierend auf Geräteklasse + Artikelname
SUBCLASS_RULES = {
    "Schutzkleidung und Schutzgerät": [
        (["helm", "visier"], "Helme"),
        (["brandbekämpfung", "nomex", "nti"], "Schutzkleidung Brandbekämpfung"),
        (["schnittschutz", "holzfäller", "forstarbeiter"], "Schutzkleidung TH"),
    ],
    "Löschgerät": [
        (["feuerlöscher"], "Tragbare Feuerlöscher"),
    ],
    "Schläuche, Armaturen und Zubehör": [
        (["schlauch", "saugschlauch", "druckschlauch"], "Schläuche"),
    ],
    "Rettungsgerät": [
        (["haltegurt", "auffanggurt"], "Feuerwehrhaltegurte"),
        (["leine", "fangleine"], "Feuerwehrleinen"),
        (["spanngurt", "seil", "schlinge", "rundschlinge"], "Spanngurte & Seile"),
        (["leiter", "steckleiter"], "Tragbare Leitern"),
    ],
    "Beleuchtungs-, Signal- und Fernmeldegerät": [
        (["funk", "melder", "meldeempfänger"], "Funkgeräte & Melder"),
        (["leitkegel", "verkehr", "absperrband"], "Geräte Verkehrssicherung"),
    ],
    "Pumpen": [
        (["pumpe", "fp "], "Pumpen"),
    ],
    "Atemschutz": [
        (["atemschutz", "pressluftatmer", "lungenautomat"], "Atemschutzgeräte"),
    ],
    "Sondergerät": [
        (["elektr", "trennschleif", "winkelschleif"], "Elektrische Geräte"),
    ],
}

# Default-Unterklasse pro Geräteklasse
SUBCLASS_DEFAULTS = {
    "Schutzkleidung und Schutzgerät": "Schutzkleidung sonstige",
    "Löschgerät": "Löschgeräte",
    "Schläuche, Armaturen und Zubehör": "Wasserführende Armaturen",
    "Rettungsgerät": "Rettungsgeräte",
    "Sanitäts- und Wiederbelebungsgerät": "Sanitäts- & Wiederbelebungsgeräte",
    "Beleuchtungs-, Signal- und Fernmeldegerät": "Signal- & Beleuchtungsgeräte",
    "Arbeitsgerät": "Geräte & Werkzeuge",
    "Handwerkzeug und Messgerät": "",
    "Sondergerät": "Elektrische Geräte",
    "Pumpen": "Pumpen",
    "Atemschutz": "Atemschutzgeräte",
}


def guess_subclass(device_class, article_name):
    if not device_class or not article_name:
        return ""
    name_lower = article_name.lower()
    rules = SUBCLASS_RULES.get(device_class, [])
    for keywords, subclass in rules:
        if any(kw in name_lower for kw in keywords):
            return subclass
    return SUBCLASS_DEFAULTS.get(device_class, "")


def parse_interval(text):
    if not text:
        return None
    t = str(text).strip().lower()
    return INTERVAL_MAP.get(t)


def parse_retirement(text):
    if not text:
        return None
    t = str(text).strip().lower()
    return RETIREMENT_MAP.get(t)


def parse_date(val):
    if not val:
        return ""
    s = str(val).strip()
    if "00:00:00" in s:
        return s.split(" ")[0]
    try:
        from datetime import datetime
        if hasattr(val, 'strftime'):
            return val.strftime('%Y-%m-%d')
    except:
        pass
    return s


def main():
    src_path = "LOKAL_Geräteliste LB 12.xlsx"
    dst_path = "FFH-LB12_Inventar_DB_v9.xlsx"

    print(f"Lese {src_path}...")
    wb_src = openpyxl.load_workbook(src_path, data_only=True)
    ws_src = wb_src[wb_src.sheetnames[0]]

    wb_dst = openpyxl.Workbook()
    ws_dst = wb_dst.active
    ws_dst.title = "Artikel"

    # DB-kompatible Spalten (CSV-Import-Format)
    headers = [
        "name",
        "inventoryNumber",
        "manufacturer",
        "articleType",
        "description",
        "inspectionInterval",
        "value",
        "serialNumber",
        "din",
        "specification",
        "manufacturingDate",
        "warehouse",
        "deviceClass",
        "deviceSubclass",
        "designationLB",
        "commissionedDate",
        "decommissionedDate",
        "communityInventoryNumber",
        "mpFeuerInventoryNumber",
        "retirementPeriodMonths",
        "pruefgrundsaetze",
    ]

    # Styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, size=10, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )

    # Write headers
    for c, h in enumerate(headers, 1):
        cell = ws_dst.cell(1, c, h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    # Process source rows
    row_out = 2
    skipped = 0
    for r in range(2, ws_src.max_row + 1):
        name = ws_src.cell(r, 3).value  # Bezeichnung LB 12
        if not name or not str(name).strip():
            skipped += 1
            continue

        name = str(name).strip()
        inv_id = str(ws_src.cell(r, 1).value or "").strip()
        mp_feuer = str(ws_src.cell(r, 2).value or "").strip()
        standort = str(ws_src.cell(r, 4).value or "").strip()
        geraeteart = str(ws_src.cell(r, 5).value or "").strip()
        seriennr = str(ws_src.cell(r, 6).value or "").strip()
        inv_gemeinde = str(ws_src.cell(r, 7).value or "").strip()
        herst_bez = str(ws_src.cell(r, 8).value or "").strip()
        hersteller = str(ws_src.cell(r, 9).value or "").strip()
        herst_datum = parse_date(ws_src.cell(r, 10).value)
        indienst = str(ws_src.cell(r, 11).value or "").strip()
        pruef_interval_sf = parse_interval(ws_src.cell(r, 18).value)
        aussonderung = parse_retirement(ws_src.cell(r, 20).value)
        ausserdienst = parse_date(ws_src.cell(r, 21).value)
        pruefgrundlagen = str(ws_src.cell(r, 22).value or "").strip()
        bemerkungen = str(ws_src.cell(r, 23).value or "").strip()

        # Normalize device class
        dc_normalized = DEVICE_CLASS_NORMALIZE.get(geraeteart.lower(), geraeteart)

        # Resolve warehouse
        warehouse = STANDORT_MAP.get(standort.lower(), standort)

        # Guess subclass
        subclass = guess_subclass(dc_normalized, name)

        # Handle Indienststellung (sometimes just a year)
        if indienst and len(indienst) == 4 and indienst.isdigit():
            indienst = f"{indienst}-01-01"
        elif "00:00:00" in indienst:
            indienst = indienst.split(" ")[0]

        # Write row
        vals = [
            name,                                     # name
            inv_id,                                   # inventoryNumber
            hersteller,                               # manufacturer
            "",                                       # articleType
            bemerkungen,                              # description
            pruef_interval_sf,                        # inspectionInterval
            "",                                       # value
            seriennr,                                 # serialNumber
            "",                                       # din
            herst_bez,                                # specification
            herst_datum,                              # manufacturingDate
            warehouse,                                # warehouse
            dc_normalized,                            # deviceClass
            subclass,                                 # deviceSubclass
            "LB12",                                   # designationLB
            indienst,                                 # commissionedDate
            ausserdienst,                             # decommissionedDate
            inv_gemeinde,                             # communityInventoryNumber
            mp_feuer,                                 # mpFeuerInventoryNumber
            aussonderung,                             # retirementPeriodMonths
            pruefgrundlagen,                          # pruefgrundsaetze
        ]

        for c, v in enumerate(vals, 1):
            if v is not None and v != "":
                ws_dst.cell(row_out, c, v)

        row_out += 1

    # Auto column widths
    for c, h in enumerate(headers, 1):
        max_len = len(h)
        for row in range(2, row_out):
            val = ws_dst.cell(row, c).value
            if val:
                max_len = max(max_len, min(len(str(val)), 40))
        ws_dst.column_dimensions[openpyxl.utils.get_column_letter(c)].width = max_len + 2

    ws_dst.freeze_panes = "A2"
    ws_dst.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(headers))}{row_out - 1}"

    # Add info sheet
    ws_info = wb_dst.create_sheet("Info")
    ws_info.cell(1, 1, "FFH-LB12 Inventar DB v9")
    ws_info.cell(2, 1, f"Erstellt aus: {src_path}")
    ws_info.cell(3, 1, f"Artikel: {row_out - 2}")
    ws_info.cell(4, 1, f"Übersprungen (leer): {skipped}")
    ws_info.cell(6, 1, "Spalten sind CSV-Import-kompatibel.")
    ws_info.cell(7, 1, "Kann direkt über Einstellungen → Datenimport importiert werden.")
    ws_info.cell(8, 1, "Trennzeichen: Semikolon (;), Kodierung: UTF-8")
    ws_info.cell(10, 1, "Hinweis: Die Spalte 'pruefgrundsaetze' wird beim Import nicht")
    ws_info.cell(11, 1, "automatisch übernommen - diese müssen manuell in der")
    ws_info.cell(12, 1, "Artikel-Detailseite unter 'Prüfgrundsätze' eingetragen werden.")

    total = row_out - 2
    print(f"\n{total} Artikel konvertiert, {skipped} leere Zeilen übersprungen")
    print(f"Speichere {dst_path}...")
    wb_dst.save(dst_path)
    print("Fertig!")


if __name__ == "__main__":
    main()
