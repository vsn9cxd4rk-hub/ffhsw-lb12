#!/usr/bin/env python3
"""
Erstellt FFH-LB12_Inventar_DB_v8.xlsx aus v7 mit folgenden Erweiterungen:
- Neue Spalten: Geräteklasse, Unterklasse
- Sheet-Name -> Geräteklasse/Unterklasse Mapping
- Bestehende Daten bleiben erhalten
- Spalten werden vereinheitlicht
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from copy import copy

# Mapping: Sheet-Name in v7 -> (Geräteklasse, Unterklasse oder None)
SHEET_CLASS_MAP = {
    "1_PSA": ("PSA", None),  # Unterklasse wird aus den Daten abgeleitet
    "2_ErsteHilfe&Hygiene": ("Erste Hilfe & Hygiene", "Sanitäts- & Wiederbelebungsgeräte"),
    "3_Signal&Beleuchtungsgeräte": ("Signal- & Beleuchtungsgeräte", None),
    "4_Arbeitsgeräte": ("Arbeitsgeräte", None),
    "5_Löschgeräte": ("Löschgeräte", None),
    "6_Rettungsgeräte": ("Rettungsgeräte", None),
    "7_Elektrische-Geräte": ("Elektrische Geräte", "Elektrische Geräte"),
    "8_Fahrzeughalle&Werkstatt": ("Geräte & Fahrzeuge im GH", "Geräte & Fahrzeuge"),
    "9_Schulungsraum&Ausbildung": ("Geräte & Fahrzeuge im GH", "Geräte & Fahrzeuge"),
    "10_Küche-Fahrzeughalle": ("Geräte & Fahrzeuge im GH", "Geräte & Fahrzeuge"),
    "11_BüroLBZF": ("Geräte & Fahrzeuge im GH", "Geräte & Fahrzeuge"),
    "12_EDV": ("Geräte & Fahrzeuge im GH", "Geräte & Fahrzeuge"),
}

# Heuristic: Unterklasse anhand Artikelname für Sheets ohne feste Zuordnung
PSA_SUBCLASS_RULES = [
    (["helm", "visier"], "Helme"),
    (["atemschutz", "pressluftatmer", "atemanschluss", "lungenautomat", "fluchthaube", "maske"], "Pressluftatmer"),
    (["brandbekämpfung", "nomex", "nti ", "feuerschutz", "hitze"], "Schutzkleidung Brandbekämpfung"),
    (["schnittschutz", "holzfäller", "th ", "technische"], "Schutzkleidung TH"),
    (["schutzbrille", "warnweste", "handschuh", "stiefel", "gummistiefel", "overall", "schutzkleidung", "beinling"], "Schutzkleidung sonstige"),
]

SIGNAL_SUBCLASS_RULES = [
    (["funk", "melder", "meldeempfänger", "digitalfunk", "handfunk"], "Funkgeräte & Melder"),
    (["leitkegel", "absperrband", "verkehr", "faltdreieck", "warndreieck"], "Geräte Verkehrssicherung"),
]

ARBEIT_SUBCLASS_RULES = [
    (["pumpe", "tauchpumpe"], "Pumpen"),
]

LOESCH_SUBCLASS_RULES = [
    (["schlauch", "druckschlauch", "saugschlauch"], "Schläuche"),
    (["feuerlöscher"], "Tragbare Feuerlöscher"),
    (["armatur", "verteiler", "sammelstück", "übergangsstück", "kupplung", "strahlrohr", "standrohr", "hydrant", "schlüssel"], "Wasserführende Armaturen"),
]

RETTUNG_SUBCLASS_RULES = [
    (["haltegurt", "sicherheitsgurt", "auffanggurt"], "Feuerwehrhaltegurte"),
    (["leine", "leinenbeutel", "fangleine"], "Feuerwehrleinen"),
    (["spanngurt", "seil", "schlinge", "bandschlinge", "zurrgurt"], "Spanngurte & Seile"),
    (["leiter", "steckleiter", "schiebeleiter", "klappleiter"], "Tragbare Leitern"),
]


def guess_subclass(sheet_name, article_name):
    """Bestimmt die Unterklasse basierend auf Sheet-Name und Artikelname."""
    if not article_name:
        return None
    name_lower = article_name.lower()

    rules_map = {
        "1_PSA": PSA_SUBCLASS_RULES,
        "3_Signal&Beleuchtungsgeräte": SIGNAL_SUBCLASS_RULES,
        "4_Arbeitsgeräte": ARBEIT_SUBCLASS_RULES,
        "5_Löschgeräte": LOESCH_SUBCLASS_RULES,
        "6_Rettungsgeräte": RETTUNG_SUBCLASS_RULES,
    }

    rules = rules_map.get(sheet_name, [])
    for keywords, subclass in rules:
        if any(kw in name_lower for kw in keywords):
            return subclass

    # Defaults per sheet if no rule matches
    defaults = {
        "1_PSA": "Schutzkleidung sonstige",
        "3_Signal&Beleuchtungsgeräte": "Signal- & Beleuchtungsgeräte",
        "4_Arbeitsgeräte": "Geräte & Werkzeuge",
        "5_Löschgeräte": "Löschgeräte",
        "6_Rettungsgeräte": "Rettungsgeräte",
    }
    return defaults.get(sheet_name)


# Vereinheitlichte Spaltenreihenfolge für v8
V8_HEADERS = [
    "Lagerort",
    "Inventarnummer",
    "Name",
    "Typ",
    "Spezifikation",
    "Beschreibung",
    "Seriennummer",
    "Hersteller",
    "Herstellerdatum",
    "Wert",
    "Bemerkung",
    "Prüfintervall [Monate]",
    "NORM [DIN]",
    "Geräteklasse",
    "Unterklasse",
]

# Header-Normalisierung: verschiedene Schreibweisen -> einheitlicher Index
HEADER_ALIASES = {
    "lagerort": 0,
    "inventarnummer": 1,
    "name": 2,
    "artikelbezeichnung": 2,
    "typ": 3,
    "spezifikation": 4,
    "beschreibung": 5,
    "seriennummer": 6,
    "hersteller": 7,
    "herstellerdatum": 8,
    "herstellerdatum [jahr]": 8,
    "wert": 9,
    "bemerkung": 10,
    "beschreibung": 10,  # manche Sheets nutzen Beschreibung statt Bemerkung
    "prüfintervall [monate]": 11,
    "prüfintervall []": 11,
    "prpfintervall [monate]": 11,
    "norm [din]": 12,
}


def main():
    src_path = "FFH-LB12_Inventar_DB_v7.xlsx"
    dst_path = "FFH-LB12_Inventar_DB_v8.xlsx"

    print(f"Lese {src_path}...")
    wb_src = openpyxl.load_workbook(src_path, data_only=True)

    wb_dst = openpyxl.Workbook()
    # Remove default sheet
    wb_dst.remove(wb_dst.active)

    # Styles
    header_font = Font(bold=True, size=10)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=10, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )
    new_col_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

    total_rows = 0

    for sheet_name in wb_src.sheetnames:
        ws_src = wb_src[sheet_name]
        ws_dst = wb_dst.create_sheet(title=sheet_name)

        class_info = SHEET_CLASS_MAP.get(sheet_name, (None, None))
        device_class = class_info[0]
        fixed_subclass = class_info[1]

        # Read source headers (row 1) and build column mapping
        src_headers = []
        for c in range(1, ws_src.max_column + 1):
            val = ws_src.cell(1, c).value
            src_headers.append(str(val).strip() if val else "")

        # Build mapping: v8 column index -> source column index
        col_map = {}
        for src_idx, hdr in enumerate(src_headers):
            hdr_lower = hdr.lower().strip()
            if hdr_lower in HEADER_ALIASES:
                v8_idx = HEADER_ALIASES[hdr_lower]
                if v8_idx not in col_map:  # first match wins
                    col_map[v8_idx] = src_idx

        # Special: some sheets have "Beschreibung" that should map to Bemerkung (col 10)
        # but also a real Beschreibung (col 5). Handle by checking if col 5 is already mapped.
        # If the sheet has both a dedicated "Beschreibung" column and another,
        # the second one maps to Bemerkung.
        desc_cols = [i for i, h in enumerate(src_headers) if h.lower().strip() in ("beschreibung", "bemerkung")]
        if len(desc_cols) >= 2:
            col_map[5] = desc_cols[0]  # First = Beschreibung
            col_map[10] = desc_cols[1]  # Second = Bemerkung
        elif len(desc_cols) == 1:
            # If sheet has 13 cols (with separate Beschreibung), put in col 5
            if ws_src.max_column >= 13 and 5 not in col_map:
                col_map[5] = desc_cols[0]
            elif 10 not in col_map:
                col_map[10] = desc_cols[0]

        # Write v8 headers
        for c, hdr in enumerate(V8_HEADERS, 1):
            cell = ws_dst.cell(1, c, hdr)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
            # Highlight new columns
            if hdr in ("Geräteklasse", "Unterklasse"):
                cell.fill = PatternFill(start_color="548235", end_color="548235", fill_type="solid")

        # Copy data rows
        sheet_rows = 0
        for r in range(2, ws_src.max_row + 1):
            # Skip empty rows
            name_col = col_map.get(2)
            if name_col is not None:
                name_val = ws_src.cell(r, name_col + 1).value
            else:
                name_val = None
            if not name_val:
                continue

            article_name = str(name_val).strip()

            for v8_col_idx in range(13):  # columns 0-12
                src_col_idx = col_map.get(v8_col_idx)
                if src_col_idx is not None:
                    val = ws_src.cell(r, src_col_idx + 1).value
                    if val is not None:
                        ws_dst.cell(sheet_rows + 2, v8_col_idx + 1, val)

            # Geräteklasse (col 14)
            cell_class = ws_dst.cell(sheet_rows + 2, 14, device_class or "")
            cell_class.fill = new_col_fill

            # Unterklasse (col 15)
            subclass = fixed_subclass or guess_subclass(sheet_name, article_name)
            cell_sub = ws_dst.cell(sheet_rows + 2, 15, subclass or "")
            cell_sub.fill = new_col_fill

            sheet_rows += 1

        total_rows += sheet_rows
        print(f"  {sheet_name}: {sheet_rows} Zeilen -> Klasse: {device_class}")

    # Adjust column widths
    for ws in wb_dst.worksheets:
        for col_idx, hdr in enumerate(V8_HEADERS, 1):
            max_len = len(hdr)
            for row in range(2, ws.max_row + 1):
                val = ws.cell(row, col_idx).value
                if val:
                    max_len = max(max_len, min(len(str(val)), 40))
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = max_len + 2

        # Freeze header row
        ws.freeze_panes = "A2"
        # Auto-filter
        ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(V8_HEADERS))}{ws.max_row}"

    print(f"\nGesamt: {total_rows} Artikel über {len(wb_dst.sheetnames)} Sheets")
    print(f"Speichere {dst_path}...")
    wb_dst.save(dst_path)
    print("Fertig!")


if __name__ == "__main__":
    main()
